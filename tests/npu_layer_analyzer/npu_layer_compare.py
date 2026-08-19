#!/usr/bin/env python3
"""NPU Layer Compare — 统一入口：一键运行 npu_layer_analyzer + layer_analyzer + layer_compare。

用法：
  # 基本用法：CSV + JSON → 对比结果
  python npu_layer_compare.py --csv kernel_details.csv --json qwen1.json

  # 指定 task-id（定位特定 forward segment）
  python npu_layer_compare.py --csv kernel_details.csv --json qwen1.json --task-id 41500

  # 指定输出目录
  python npu_layer_compare.py --csv kernel_details.csv --json qwen1.json -o my_compare

  # 只跑 npu_layer_analyzer（不跑 layer_analyzer）
  python npu_layer_compare.py --csv kernel_details.csv --npu-only

  # 只跑 layer_analyzer（不跑 npu_layer_analyzer）
  python npu_layer_compare.py --json qwen1.json --layer-only

输出结构（全部 xlsx）：
  <output_dir>/
  ├── npu_out.xlsx                # npu_layer_analyzer 输出（多 Sheet）
  ├── layer_out.xlsx              # layer_analyzer 输出（多 Sheet）
  └── compare_result.xlsx         # 对比结果（总比较 + 算子明细）
"""

from __future__ import annotations

import argparse
import csv
import shutil
import subprocess
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SCRIPT_DIR = Path(__file__).resolve().parent

# Named constants (replacing magic numbers)
EXCEL_SHEET_NAME_MAX_LEN = 31
COL_WIDTH_SAMPLE_ROWS = 100
COL_WIDTH_PADDING = 4
COL_MAX_WIDTH = 50

# 文件名前缀（与 npu_layer_analyzer.py / layer_analyzer.py 一致）
NPU_OUTPUT_PREFIX = "真实"
LAYER_OUTPUT_PREFIX = "仿真"


def run_cmd(cmd: list[str], desc: str) -> int:
    """运行子命令，实时打印输出。"""
    print(f"\n{'=' * 70}")
    print(f"[STEP] {desc}")
    print(f"  CMD: {' '.join(str(c) for c in cmd)}")
    print(f"{'=' * 70}")
    result = subprocess.run(cmd, cwd=str(SCRIPT_DIR))
    if result.returncode != 0:
        print(f"[ERROR] {desc} 失败 (exit code={result.returncode})")
    return result.returncode


# ── CSV → XLSX 转换 ──────────────────────────────────────────────────────

_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill("solid", fgColor="4472C4")
_THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def _sheet_name_from_csv(csv_path: Path, max_len: int = EXCEL_SHEET_NAME_MAX_LEN) -> str:
    """从 CSV 文件名生成 Sheet 名（Excel 限制 31 字符）。"""
    name = csv_path.stem
    # 截断到 max_len
    return name[:max_len]


def csv_to_sheet(ws, csv_path: Path) -> None:
    """读取 CSV 写入 worksheet，带表头样式和自动列宽。"""
    with csv_path.open("r", encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        rows = list(reader)

    if not rows:
        return

    # 写入数据
    for r_idx, row in enumerate(rows, start=1):
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = _THIN_BORDER
            if r_idx == 1:
                cell.font = _HEADER_FONT
                cell.fill = _HEADER_FILL
                cell.alignment = Alignment(horizontal="center")

    # 自动列宽（采样前 100 行）
    for c_idx in range(1, len(rows[0]) + 1):
        max_len = 0
        for r_idx in range(1, min(len(rows), COL_WIDTH_SAMPLE_ROWS) + 1):
            val = ws.cell(row=r_idx, column=c_idx).value
            if val:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[get_column_letter(c_idx)].width = min(max_len + COL_WIDTH_PADDING, COL_MAX_WIDTH)

    # 冻结首行
    ws.freeze_panes = "A2"


def csv_dir_to_xlsx(csv_dir: Path, output_xlsx: Path, exclude_patterns: list[str] | None = None) -> bool:
    """把目录下所有 CSV 合并到一个 xlsx（每个 CSV 一个 Sheet），然后删除 CSV 和目录。

    exclude_patterns: 文件名包含这些模式的 CSV 不纳入 xlsx（如中间转换文件）。
    返回 True 如果成功生成 xlsx。
    """
    csv_files = sorted(csv_dir.glob("*.csv"))
    if not csv_files:
        return False

    # 过滤掉不需要的中间文件
    if exclude_patterns:
        csv_files = [f for f in csv_files if not any(p in f.name for p in exclude_patterns)]

    if not csv_files:
        return False

    wb = Workbook()
    # 删除默认 Sheet
    wb.remove(wb.active)

    for csv_file in csv_files:
        sheet_name = _sheet_name_from_csv(csv_file)
        ws = wb.create_sheet(title=sheet_name)
        csv_to_sheet(ws, csv_file)

    wb.save(str(output_xlsx))

    # 删除原 CSV 目录
    shutil.rmtree(csv_dir, ignore_errors=True)

    sheet_names = [ws.title for ws in wb.worksheets]
    print(f"  → {output_xlsx.name} ({len(sheet_names)} sheets: {', '.join(sheet_names)})")
    return True


def collect_layer_csvs(directory: Path, prefix: str) -> dict[str, Path]:
    """收集目录下所有层 CSV，按类型分类。

    返回 {"dense": path, "moe": path} 或 {"dense": path}（无 MoE 时）。
    文件名规则（层号在文件名中）：
      {prefix}..._layer{N}_dense.csv → Dense（MoE 模型）
      {prefix}..._layer{N}_moe.csv   → MoE
      {prefix}..._layer{N}.csv       → Dense（非 MoE 模型，无 _dense/_moe 后缀）
    """
    result: dict[str, Path] = {}

    # MoE 模型：_layer{N}_dense.csv 和 _layer{N}_moe.csv
    dense_files = sorted(directory.glob(f"{prefix}*_layer*_dense.csv"))
    if dense_files:
        result["dense"] = dense_files[-1]

    moe_files = sorted(directory.glob(f"{prefix}*_layer*_moe.csv"))
    if moe_files:
        result["moe"] = moe_files[-1]

    # 非 MoE 模型：_layer{N}.csv（排除 _dense/_moe/_layered）
    if "dense" not in result:
        plain_files = [
            f
            for f in sorted(directory.glob(f"{prefix}*_layer*.csv"))
            if "_dense" not in f.name and "_moe" not in f.name and "_layered" not in f.name
        ]
        if plain_files:
            result["dense"] = plain_files[-1]

    return result


def merge_compare_xlsx(
    pairs: list[tuple[str, Path]],
    output_path: Path,
) -> None:
    """将多个 layer_compare xlsx 合并为一个，sheet 名加类型前缀。

    pairs: [(layer_type, xlsx_path), ...]
    """
    from openpyxl import load_workbook

    merged_wb = Workbook()
    merged_wb.remove(merged_wb.active)

    for layer_type, xlsx_path in pairs:
        prefix = f"{'Dense' if layer_type == 'dense' else 'MoE'}_"
        src_wb = load_workbook(str(xlsx_path))
        for src_ws in src_wb.worksheets:
            new_name = f"{prefix}{src_ws.title}"[:EXCEL_SHEET_NAME_MAX_LEN]
            dst_ws = merged_wb.create_sheet(title=new_name)
            # 复制单元格（含样式）
            for row in src_ws.iter_rows():
                for cell in row:
                    dst_cell = dst_ws.cell(row=cell.row, column=cell.column, value=cell.value)
                    if cell.has_style:
                        dst_cell.font = cell.font.copy()
                        dst_cell.fill = cell.fill.copy()
                        dst_cell.alignment = cell.alignment.copy()
                        dst_cell.border = cell.border.copy()
            # 复制列宽
            for col_letter, dim in src_ws.column_dimensions.items():
                if dim.width:
                    dst_ws.column_dimensions[col_letter].width = dim.width
            # 复制冻结窗格
            if src_ws.freeze_panes:
                dst_ws.freeze_panes = src_ws.freeze_panes

    merged_wb.save(str(output_path))


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(
        description="NPU Layer Compare — 统一入口：npu_layer_analyzer + layer_analyzer + layer_compare"
    )
    parser.add_argument("--csv", default=None, help="输入 CSV 文件（给 npu_layer_analyzer）")
    parser.add_argument(
        "--json",
        default=None,
        help="输入 JSON 或 CSV 文件（给 layer_analyzer；JSON 自动转 CSV）",
    )
    parser.add_argument(
        "-o",
        "--output-dir",
        default="compare_test",
        help="输出目录（默认 compare_test）",
    )
    parser.add_argument(
        "--task-id",
        type=int,
        default=None,
        help="算子 Task ID，定位特定 forward segment",
    )
    parser.add_argument("--npu-only", action="store_true", help="只跑 npu_layer_analyzer")
    parser.add_argument("--layer-only", action="store_true", help="只跑 layer_analyzer")
    parser.add_argument("--no-compare", action="store_true", help="不跑 layer_compare")
    parser.add_argument("--layer-index", type=int, default=None, help="指定 Dense 层号")
    args = parser.parse_args(argv if argv is not None else sys.argv[1:])

    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    npu_layer_csvs: dict[str, Path] = {}  # type → path (npu_layer_analyzer 层提取结果)
    sim_layer_csvs: dict[str, Path] = {}  # type → path (layer_analyzer 层提取结果)

    # ── Step 1: npu_layer_analyzer ──
    if not args.layer_only and args.csv:
        csv_path = Path(args.csv)
        if not csv_path.is_file():
            print(f"[ERROR] CSV 不存在: {csv_path}")
            return 1

        npu_out = output_dir / "npu_out"
        npu_cmd = [
            sys.executable,
            str(SCRIPT_DIR / "npu_layer_analyzer.py"),
            "--input",
            str(csv_path),
            "--output-dir",
            str(npu_out),
        ]
        if args.task_id is not None:
            npu_cmd.extend(["--task-id", str(args.task_id)])

        rc = run_cmd(npu_cmd, "npu_layer_analyzer (CSV → Forward 切分 + 层提取)")
        if rc != 0:
            return rc
        npu_layer_csvs = collect_layer_csvs(npu_out, NPU_OUTPUT_PREFIX)
        for layer_type, path in npu_layer_csvs.items():
            print(f"  → 层提取结果 ({layer_type}): {path}")

    # ── Step 2: layer_analyzer ──
    if not args.npu_only and args.json:
        json_path = Path(args.json)
        if not json_path.is_file():
            print(f"[ERROR] 文件不存在: {json_path}")
            return 1

        layer_out = output_dir / "layer_out"
        layer_out.mkdir(parents=True, exist_ok=True)

        # 2a: JSON → CSV（如果输入是 JSON），否则直接用 CSV
        if json_path.suffix.lower() == ".json":
            intermediate_csv = layer_out / f"{json_path.stem}_kernel_details.csv"
            convert_cmd = [
                sys.executable,
                str(SCRIPT_DIR / "trace_json_to_csv.py"),
                "--input",
                str(json_path),
                "--output",
                str(intermediate_csv),
            ]
            rc = run_cmd(convert_cmd, "trace_json_to_csv (JSON → CSV)")
            if rc != 0:
                print("[ERROR] JSON → CSV 转换失败")
                return 1
            layer_input_csv = intermediate_csv
        else:
            # 已经是 CSV，直接用
            layer_input_csv = json_path

        # 2b: layer_analyzer
        output_prefix = layer_out / layer_input_csv.stem
        layer_cmd = [
            sys.executable,
            str(SCRIPT_DIR / "layer_analyzer.py"),
            "--input",
            str(layer_input_csv),
            "--output",
            str(output_prefix),
            "--delimiter",
            "attention",
        ]
        if args.layer_index is not None:
            layer_cmd.extend(["--layer-index", str(args.layer_index)])

        rc = run_cmd(layer_cmd, "layer_analyzer (CSV → 层标注 + 层提取)")
        if rc != 0:
            return rc
        # layer_analyzer 输出到 --output 指定的目录（layer_out/）
        # 输出文件名带"仿真"前缀（由 layer_analyzer.py 内部添加）
        base = output_prefix.stem
        sim_base = f"{LAYER_OUTPUT_PREFIX}{base}"
        for suffix in [
            "_layer_dense.csv",
            "_layer_moe.csv",
            "_layer.csv",
            "_layered.csv",
        ]:
            src_file = layer_out / f"{sim_base}{suffix}"
            # 如果 layered CSV 还在临时目录（兼容旧逻辑），移动过来
            old_src = layer_input_csv.parent / f"{sim_base}{suffix}"
            if old_src.is_file() and old_src != src_file:
                shutil.move(str(old_src), str(src_file))
        sim_layer_csvs = collect_layer_csvs(layer_out, LAYER_OUTPUT_PREFIX)
        for layer_type, path in sim_layer_csvs.items():
            print(f"  → 层提取结果 ({layer_type}): {path}")

    # ── Step 3: layer_compare（自动配对 Dense↔Dense, MoE↔MoE）──
    if not args.no_compare:
        # 按类型配对
        pairs = []
        for layer_type in ["dense", "moe"]:
            npu_csv = npu_layer_csvs.get(layer_type)
            sim_csv = sim_layer_csvs.get(layer_type)
            if npu_csv and sim_csv:
                pairs.append((layer_type, npu_csv, sim_csv))

        if pairs:
            compare_output = output_dir / "compare_result.xlsx"
            done_pairs: list[tuple[str, Path]] = []

            for layer_type, npu_csv, sim_csv in pairs:
                label = "Dense" if layer_type == "dense" else "MoE"
                if len(pairs) == 1:
                    temp_output = compare_output
                else:
                    temp_output = output_dir / f"compare_{layer_type}.xlsx"

                compare_cmd = [
                    sys.executable,
                    str(SCRIPT_DIR / "layer_compare.py"),
                    "-a",
                    str(npu_csv),
                    "-b",
                    str(sim_csv),
                    "-o",
                    str(temp_output),
                ]
                rc = run_cmd(compare_cmd, f"layer_compare ({label} 层对比 → xlsx)")
                if rc != 0:
                    return rc
                done_pairs.append((layer_type, temp_output))

            # 合并多个 xlsx 为一个
            if len(done_pairs) > 1:
                merge_compare_xlsx(done_pairs, compare_output)
                # 删除临时文件
                for _, temp_path in done_pairs:
                    if temp_path != compare_output:
                        temp_path.unlink(missing_ok=True)

            if done_pairs:
                print(f"\n{'=' * 70}")
                print(f"[DONE] 对比结果: {compare_output}")
                for layer_type, _ in done_pairs:
                    label = "Dense" if layer_type == "dense" else "MoE"
                    prefix = f"{label}_" if len(done_pairs) > 1 else ""
                    print(f"  {prefix}总比较（按 Stage 汇总时间）")
                    print(f"  {prefix}算子明细（逐算子并排对比）")
                print(f"{'=' * 70}")
        else:
            missing = []
            if not npu_layer_csvs:
                missing.append("npu_layer_analyzer 层提取结果")
            if not sim_layer_csvs:
                missing.append("layer_analyzer 层提取结果")
            print(f"\n[SKIP] layer_compare：缺少 {', '.join(missing)}")

    # ── Step 4: CSV → XLSX 统一格式 ──
    print(f"\n{'=' * 70}")
    print("[STEP] CSV → XLSX 统一格式")
    print(f"{'=' * 70}")

    npu_out_dir = output_dir / "npu_out"
    if npu_out_dir.is_dir():
        npu_xlsx = output_dir / "npu_out.xlsx"
        csv_dir_to_xlsx(npu_out_dir, npu_xlsx)

    layer_out_dir = output_dir / "layer_out"
    if layer_out_dir.is_dir():
        layer_xlsx = output_dir / "layer_out.xlsx"
        # 排除 JSON→CSV 的中间文件（无标注，和 layered 重复）
        csv_dir_to_xlsx(layer_out_dir, layer_xlsx, exclude_patterns=["_kernel_details.csv"])

    print(f"\n{'=' * 70}")
    print(f"[DONE] 全部完成，输出目录: {output_dir.resolve()}")
    for xlsx in sorted(output_dir.glob("*.xlsx")):
        print(f"  → {xlsx.name}")
    print(f"{'=' * 70}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
